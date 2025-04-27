import streamlit as st
import pandas as pd
import smtplib
import datetime
import networkx as nx
import matplotlib.pyplot as plt
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from io import BytesIO
import base64
from graphviz import Digraph
from faker import Faker
import random
import re

# --- Email Config ---
EMAIL_CONFIG = {
    'smtp_server': 'smtp.gmail.com',
    'smtp_port': 587,
    'sender_email': 'auditflow5@gmail.com',
    'sender_password': st.secrets.get('EMAIL_PASSWORD', '')
}

# --- Custom Background + Button Style ---
st.markdown("""
    <style>
        .stApp {
            background: linear-gradient(135deg, #ffd6e0, #dbeafe);
            background-attachment: fixed;
            background-size: cover;
            color: #000000;
        }
        .card {
            background-color: rgba(255, 255, 255, 0.9);
            padding: 20px;
            border-radius: 12px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            margin-bottom: 20px;
        }
        .stButton > button {
            background-color: #0066cc;
            color: white;
            padding: 0.5em 1.2em;
            border: none;
            border-radius: 8px;
            transition: 0.3s;
        }
        .stButton > button:hover {
            background-color: #004c99;
            transform: scale(1.03);
        }
    </style>
""", unsafe_allow_html=True)

# --- Main Class ---
class AuditWorkflow:
    def __init__(self):
        self.tasks = []
        self.task_graph = nx.DiGraph()
        self.faker = Faker()

    def validate_email(self, email):
        return re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email)

    def add_task(self, task_id, description, due_date, dependencies, assignee_email):
        try:
            if not task_id or task_id in [t['id'] for t in self.tasks]:
                raise ValueError("Task ID must be unique and non-empty")
            if not self.validate_email(assignee_email):
                raise ValueError("Invalid email format")
            due_date_obj = datetime.datetime.strptime(due_date, '%Y-%m-%d')
            if due_date_obj < datetime.datetime.now():
                raise ValueError("Due date cannot be in the past")
            task = {
                'id': task_id,
                'description': description,
                'due_date': due_date_obj,
                'dependencies': dependencies,
                'assignee_email': assignee_email,
                'status': 'Pending'
            }
            self.tasks.append(task)
            self.task_graph.add_node(task_id, label=description)
            for dep in dependencies:
                if dep not in [t['id'] for t in self.tasks[:-1]]:
                    raise ValueError(f"Dependency {dep} does not exist")
                self.task_graph.add_edge(dep, task_id)
            return True
        except Exception as e:
            st.error(f"Error: {str(e)}")
            return False

    def generate_fake_tasks(self, num_tasks=5):
        try:
            task_ids = [f"T{i+1}" for i in range(len(self.tasks) + 1, len(self.tasks) + num_tasks + 1)]
            for i in range(num_tasks):
                task_id = task_ids[i]
                description = self.faker.sentence(nb_words=6)
                due_date = (datetime.datetime.now() + datetime.timedelta(days=random.randint(2, 10))).strftime('%Y-%m-%d')
                existing_ids = [t['id'] for t in self.tasks] + task_ids[:i]
                dependencies = random.sample(existing_ids, min(len(existing_ids), random.randint(0, 2)))
                email = self.faker.email()
                self.add_task(task_id, description, due_date, dependencies, email)
        except Exception as e:
            st.error(f"Fake task error: {str(e)}")

    def send_reminder(self, task):
        try:
            days_left = (task['due_date'] - datetime.datetime.now()).days
            if days_left <= 2 and task['status'] == 'Pending':
                msg = MIMEMultipart()
                msg['From'] = EMAIL_CONFIG['sender_email']
                msg['To'] = task['assignee_email']
                msg['Subject'] = f"Reminder: Task '{task['description']}'"
                msg.attach(MIMEText(
                    f"Hi,\n\nThis is a reminder that the task \"{task['description']}\" is due on {task['due_date'].strftime('%Y-%m-%d')}.\n\nThank you!", 'plain'))
                with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port']) as server:
                    server.starttls()
                    server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
                    server.send_message(msg)
                return True
        except Exception as e:
            st.error(f"Email failed: {str(e)}")
        return False

    def generate_report(self):
        df = pd.DataFrame(self.tasks)
        if df.empty:
            return None
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Tasks"
        headers = ['ID', 'Description', 'Due Date', 'Dependencies', 'Assignee', 'Status']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = header
        for row, task in enumerate(self.tasks, 2):
            ws.cell(row=row, column=1).value = task['id']
            ws.cell(row=row, column=2).value = task['description']
            ws.cell(row=row, column=3).value = str(task['due_date'].date())
            ws.cell(row=row, column=4).value = ", ".join(task['dependencies'])
            ws.cell(row=row, column=5).value = task['assignee_email']
            ws.cell(row=row, column=6).value = task['status']
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def visualize_workflow(self):
        if not self.task_graph.nodes:
            return None
        dot = Digraph()
        for node in self.task_graph.nodes:
            dot.node(node, self.task_graph.nodes[node]['label'])
        for edge in self.task_graph.edges:
            dot.edge(edge[0], edge[1])
        return dot

# --- Streamlit App ---
def main():
    st.title("📋 Audit Workflow Automation")
    if 'workflow' not in st.session_state:
        st.session_state.workflow = AuditWorkflow()
    wf = st.session_state.workflow

    with st.sidebar:
        st.header("🧩 Task Input")
        task_id = st.text_input("Task ID")
        description = st.text_input("Description")
        due_date = st.text_input("Due Date (YYYY-MM-DD)")
        dependencies = st.text_input("Dependencies (comma-separated)")
        assignee = st.text_input("Assignee Email")

        if st.button("➕ Add Task"):
            dep_list = [d.strip() for d in dependencies.split(',') if d.strip()]
            if wf.add_task(task_id, description, due_date, dep_list, assignee):
                st.success("Task added successfully.")

        if st.button("✨ Generate Fake Tasks"):
            wf.generate_fake_tasks()
            st.success("Fake tasks added.")

    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.subheader("📂 Current Tasks")
    if wf.tasks:
        st.dataframe(pd.DataFrame(wf.tasks))
        if st.button("📧 Send Reminders"):
            count = sum([wf.send_reminder(t) for t in wf.tasks])
            st.success(f"Sent {count} reminder(s).")
        if st.button("📁 Download Report"):
            report = wf.generate_report()
            if report:
                b64 = base64.b64encode(report.read()).decode()
                st.markdown(f'<a href="data:application/octet-stream;base64,{b64}" download="audit_report.xlsx">📥 Click to Download Excel</a>', unsafe_allow_html=True)
    else:
        st.info("No tasks yet.")

    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.subheader("📌 Workflow Diagram")
    dot = wf.visualize_workflow()
    if dot:
        st.graphviz_chart(dot.source)
    else:
        st.warning("Add tasks to see the workflow.")
    st.markdown("</div>", unsafe_allow_html=True)

if __name__ == '__main__':
    main()
